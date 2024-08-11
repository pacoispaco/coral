#!/usr/bin/env python3

"""This module contains methods and classes for reading IOC files."""

import json
import os, os.path
import openpyxl.reader.excel as xlxs
import re

# Constants: top taxa names in IOC
TOP_TAXA_NAMES = ["NEOAVES", "NEOGNATHAE", "PALEOGNATHAE"]


class InvalidIocMasterFile (Exception):
    pass


class InvalidIocOtherListsFile (Exception):
    pass


class InvalidIocMultilingualFile (Exception):
    pass


class InvalidIocComplementaryFile (Exception):
    pass


class UnrecognizedIocFile (Exception):
    pass


class MultipleIocFilesOfSameType (Exception):
    pass


class MultipleVersionsOfIocFiles (Exception):
    pass


def _is_master_wb(wb):
    """True if 'wb' is an IOC Masterfile Excel workbook, otherwise False."""
    return wb.worksheets[0].title == "Master"


def _is_other_lists_wb(wb):
    """True if 'wb' is an IOC Other Lists Excel workbook, otherwise False."""
    return "vs_other_lists" in wb.worksheets[0].title


def _is_multilingual_wb(wb):
    """True if 'wb' is an IOC Multilingual Excel workbook, otherwise False."""
    return wb.worksheets[0].title == "List" and wb.worksheets[1].title == "Sources"


def _is_complementary_wb(wb):
    """True if 'wb' is an IOC Complementary Excel workbook, otherwise False."""
    return "IOC" in wb.worksheets[0].title


def sorted_ioc_files(filepaths):
    """Returns a list of IOC File class instances sorted in the order they
       should be processed, based on the list of 'filepaths'."""
    assert type(filepaths) is list
    result = []
    for path in filepaths:
        try:
            wb = xlxs.load_workbook(path)
        except xlxs.InvalidFileException:
            print("Error: '%s' is not an Excel file (.xlxs)." % (path))
            raise
        if _is_master_wb(wb):
            result.append(IocMasterFile(wb, path))
        elif _is_other_lists_wb(wb):
            result.append(IocOtherListsFile(wb, path))
        elif _is_multilingual_wb(wb):
            result.append(IocMultilingualFile(wb, path))
        elif _is_complementary_wb(wb):
            result.append(IocComplementaryFile(wb, path))
        else:
            msg = """Error: '%s' is not an recognized IOC file.
An IOC Master file must have the title 'Master' in the first worksheet.
An IOC Other Lists file must have the word 'vs_other_lists' in the title of the first worksheet.
An IOC Multilingual file must have the word 'List' in the title of the first worksheet and
'Sources' in the title of the second worksheet. An IOC Complementary file must have the word
'IOC' in the first worksheet."""
            print(msg % (path))
            raise UnrecognizedIocFile(path)
    # Make sure we don't have multiple IOC files of the same type
    orders = [file.order for file in result]
    if not len(set(orders)) == len(orders):
        print("Error: Multiple IOC files of same type.")
        raise MultipleIocFilesOfSameType(filepaths)
    # Make sure all IOC files have the same version
    versions = [file.version for file in result]
    if not len(set(versions)) <= 1:
        print("Error: Multiple versions of IOC files.")
        raise MultipleVersionsOfIocFiles(filepaths)
    # Return the list of IOC files sorted by their 'order' attribute
    return sorted(result, key=lambda iocfile: iocfile.order)


class IocMasterFile (object):
    """Represents a IOC Master file (Excel)."""

    def __init__(self, workbook, path):
        """Initialize with Excel 'workbook'."""
        self.order = 1
        self.workbook = workbook
        self.path = path
        self.version = None
        self.taxonomy = []        # This list contains the list of top level
                                  # taxa. Each taxa has a 'subtaxa' attribute
                                  # which is a list of taxa, etc.
        self.index = {}           # This index contains all taxa keyed by name
        self.taxonomy_stats = {}
        if self.workbook.worksheets[0].title == "Master":
            s = self.workbook.worksheets[0].cell(row=1, column=2).value
            regexp = re.compile (".*\((.*)\).*")
            self.version = regexp.match (s).groups()[0]
        else:
            print("Error; '%s' is not a avalid IOC Master File" % (self.path))
            raise InvalidIocMasterFile(self.path)

    def read(self):
        """Read the taxonomy data into the attribute 'self.taxonomy' and
           save statistics in the attribute 'self.taxonomy_stats'."""
        self.taxonomy_stats = {'infraclass_count': 0,
                               'order_count': 0,
                               'family_count': 0,
                               'genus_count': 0,
                               'species_count': 0,
                               'subspecies_count': 0}
        ws = self.workbook.worksheets[0]
        for row in ws.iter_rows(min_row=5):
            infraclass = row[0].value
            order = row[1].value
            family = row[2].value
            family_en = row[3].value
            genus = row[4].value
            species = row[5].value
            subspecies = row[6].value
            taxon = {'other_classifications': [],  # Taxa in other lists which
                                                   # are equivalent to this taxon
                     'authority': row[7].value,
                     'common_names': {'en': row[8].value},
                     'breeding_range': row[9].value,
                     'breeding_subranges': row[10].value,
                     'nonbreeding_range': row[11].value,
                     'code': row[12].value,
                     'comment': row[13].value,
                     'subtaxa': []}
            if infraclass:
                taxon['rank'] = "Infraclass"
                taxon['name'] = infraclass
                self.taxonomy.append(taxon)
                self.index[infraclass] = taxon
                self.taxonomy_stats['infraclass_count'] += 1
            elif order:
                taxon['rank'] = "Order"
                taxon['name'] = order
                taxon['supertaxon'] = self.taxonomy[-1]['name']
                self.taxonomy[-1]['subtaxa'].append(taxon)
                self.index[order] = taxon
                self.taxonomy_stats['order_count'] += 1
            elif family:
                taxon['rank'] = "Family"
                taxon['name'] = family
                taxon['supertaxon'] = self.taxonomy[-1]['subtaxa'][-1]['name']
                self.taxonomy[-1]['subtaxa'][-1]['subtaxa'].append(taxon)
                self.index[family] = taxon
                self.taxonomy_stats['family_count'] += 1
            elif genus:
                taxon['rank'] = "Genus"
                # Strip trailing "extinct" characters '\u2020' and whitespace
                taxon['name'] = genus.title().strip('\u2020').strip()
                taxon['supertaxon'] = self.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['name']
                self.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'].append(taxon)
                self.index[genus] = taxon
                self.taxonomy_stats['genus_count'] += 1
            elif species:
                taxon['rank'] = "Species"
                taxon['name'] = species
                taxon['supertaxon'] = self.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]['name']
                genus = taxon['supertaxon']
                # Strip trailing "extinct" characters '\u2020' and whitespace
                taxon['binomial_name'] = (genus + " " + species).strip('\u2020').strip()
                self.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'].append(taxon)
                self.index[taxon['binomial_name']] = taxon
                self.taxonomy_stats['species_count'] += 1
            elif subspecies:
                taxon['rank'] = "Subspecies"
                taxon['name'] = subspecies
                taxon['supertaxon'] = self.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]['name']
                genus = self.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]['name']
                species = taxon['supertaxon']
                # Strip trailing "extinct" characters '\u2020' and whitespace
                trinomial_name = (genus + " " + species + " " + subspecies).strip('\u2020').strip ()
                # Strip "extinct" characters '\u2020' in trinomial name
                taxon['trinomial_name'] = trinomial_name.replace(" \u2020", "")
                self.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'].append(taxon)
                self.index[taxon['trinomial_name']] = taxon
                self.taxonomy_stats['subspecies_count'] += 1


class IocOtherListsFile (object):
    """Represents a IOC Other Lists file (Excel)."""

    def __init__(self, workbook, path):
        """Initialize with the Excel 'workbook'."""
        self.order = 2
        self.workbook = workbook
        self.path = path
        self.version = None
        self.taxonomy = {}       # This object contains taxa indexed by their
                                 # IOC name (or binomial or trinomial name if
                                 # applicable (for species and subspecies)
        self.nonindexed = []     # List of all taxa that do not exist in IOC
                                 # but do exist in other lists
        self.taxonomy_stats = {}
        self.lists = {'ioc_7_3': 'Gill, F & D Donsker (Eds). 2017. IOC World Bird List (v 7.3)',
                      'clements_2016': 'Clements, J. F., T. S. Schulenberg, M. J. Iliff, D. Roberson, T. A. Fredericks, B. L. Sullivan, and C. L. Wood. 2016. The eBird/Clements checklist of birds of the world: v2016.',
                      'hbwbl_2016': 'del Hoyo, J., N. J. Collar, D. A. Christie, A. Elliott,  L. D. C. Fishpool, P. Boesman & G. M. Kirwan. 2014-2016. HBW and BirdLife International Illustrated Checklist of the Birds of the World, Volume 1, 2, Lynx Edicions in association with BirdLife International, Barcelona, Spain and Cambridge, UK.',
                      'hm4_4ed': 'Dickinson, E.C., J.V. Remsen Jr. & L. Christidis (Eds). 2013-2014. The Howard & Moore Complete Checklist of the Birds of the World. 4th. Edition, Vol. 1, 2, Aves Press, Eastbourne, U.K. Plus CD content: Tyrberg, T. & E.C. Dickinson Appendix 5: Extinct species. (Also includes Errata and Corrigenda to Volume 1 and List of Errata for Vol. 2 plus Corrigenda in respect of range statements and additional Errata from Vol. 1)',
                      'hbw_2013': 'del Hoyo, J., A. Elliott, J. Sargatal & D. A. Christie (Eds). 1992-2013. Handbook of the Birds of the World, Lynx Edicions.',
                      'peters_1986': 'Peters, J.L. et al. Check-list of Birds of the World, 1931-1986. Harvard University Press/Museum of Comparative Zoology.',
                      'boyd_3_08': 'John H. Boyd III - TiF checklist, Version 3.08: May 1 2017 and updated July 12 2017',
                      'hbwbl_2017': 'BirdLife International (2017) Handbook of the Birds of the World and BirdLife International digital checklist of the birds of the world. Version 9.1. Available at: http://datazone.birdlife.org/userfiles/file/Species/Taxonomy/BirdLife_Checklist_Version_91.zip [.xls zipped 1 MB].',
                      'sibley_1993': 'Sibley, C. G. and B. L. Monroe. 1993. A Supplement to Distribution and Taxonomy of Birds of the World. Yale University Press, New Haven, Connecticut.',
                      'ioc_7_2': 'Gill, F & D Donsker (Eds). 2017. IOC World Bird List (v 7.2)',
                      'ioc_7_1': 'Gill, F & D Donsker (Eds). 2017. IOC World Bird List (v 7.1)'}
        if "vs_other_lists" in self.workbook.worksheets[0].title:
            s = self.workbook.worksheets[0].cell(row=1, column=2).value
            regexp = re.compile(".*\(v (.*)\).*")
            self.version = regexp.match(s).groups()[0]
        else:
            print("Error; '%s' is not a valid IOC Other Lists File" % (self.path))
            raise InvalidIocOtherListsFile(self.path)

    def read(self):
        """Read the taxonomy data into the attribute 'self.taxonomy' and
           save statistics in the attribute 'self.taxonomy_stats'."""
        # Note that the global variable 'master_taxonomy' is a list that will maintain the
        # ordering of taxa in the IOC Master file.
        self.taxonomy_stats = {'infraclass_count': 0,
                               'order_count': 0,
                               'family_count': 0,
                               'genus_count': 0,
                               'species_count': 0,
                               'subspecies_count': 0,
                               'only_in_other_lists_count': 0}
        ws = self.workbook.worksheets[0]
        for row in ws.iter_rows(min_row=2):
            name = row[1].value
            entry = {'seq_no': row[0].value,
                     'name': name,
                     'rank': row[2].value,
                     'notes': row[3].value,
                     'iucn_red_list_category': row[32].value,
                     'following_entries': [],
                     'lists': {'clements_2016': {'name': row[5].value,
                                                 'group': row[4].value,
                                                 'family': row[18].value},
                               'hbwbl_2016': {'name': row[7].value,
                                              'group': row[6].value,
                                              'family': row[19].value},
                               'hm4_4ed': {'name': row[9].value,
                                           'group': row[8].value,
                                           'family': row[20].value},
                               'hbw_2013': {'name': row[10].value,
                                            'group': None,
                                            'family': row[21].value},
                               'peters_1986': {'name': row[11].value,
                                               'group': None,
                                               'family': row[22].value},
                               'boyd': {'name': row[12].value,
                                        'group': None,
                                        'family': row[23].value},
                               'hbwbl_2017': {'name': row[13].value,
                                              'group': None,
                                              'family': row[24].value},
                               'sibley_1993': {'name': row[14].value,
                                               'group': None,
                                               'family': row[25].value},
                               'ioc_7_2': {'name': row[15].value,
                                           'group': None,
                                           'family': row[26].value},
                               'ioc_7_1': {'name': row[16].value,
                                           'group': None,
                                           'family': row[27].value}}}
            if name:
                self.taxonomy[name] = entry
                if len(name.split()) == 2:
                    self.taxonomy_stats['species_count'] += 1
                else:
                    self.taxonomy_stats['subspecies_count'] += 1
                latest_name = name
            else:
                self.nonindexed.append(entry)
                self.taxonomy[latest_name]['following_entries'].append(entry)
                self.taxonomy_stats['only_in_other_lists_count'] += 1


class IocMultilingualFile (object):
    """Represents a IOC Multilingual file (Excel). Languages are encoded with ISO
       639-2 codes (which are not used in the actual file)."""

    def __init__(self, workbook, path):
        """Initialize with the Excel 'workbook'."""
        self.order = 3
        self.workbook = workbook
        self.path = path
        self.version = None
        self.taxonomy = {}          # This object contains taxa indexed by their
                                    # name
        self.taxonomy_stats = {}
        if self.workbook.worksheets[0].title == "List" and self.workbook.worksheets[1].title == "Sources":
            s = self.workbook.worksheets[1].cell(row=1, column=1).value
            regexp = re.compile("[A-Za-z ]*(\d*\.\d*).*")
            self.version = regexp.match(s).groups()[0]
        else:
            print("Error; '%s' is not a valid IOC Multilingual File" % (self.path))
            raise InvalidIocMultilingualFile(self.path)

    def read(self):
        """Read the taxonomy data into the attribute 'self.taxonomy' and
           save statistics in the attribute 'self.taxonomy_stats'."""
        # Note that the global variable 'master_taxonomy' is a list that will maintain the
        # ordering of taxa in the IOC Master file.
        self.taxonomy_stats = {'infraclass_count': 0,
                               'order_count': 0,
                               'family_count': 0,
                               'genus_count': 0,
                               'species_count': 0,
                               'subspecies_count': 0}
        ws = self.workbook.worksheets[0]
        i = 0
        for row in ws.iter_rows(min_row=4):
            if row[3].value and len(row[3].value.split()) == 2:
                name = row[3].value
                self.taxonomy_stats['species_count'] += 1
                i = 1
                entry = {'cat': row[6].value,
                         'cze': row[9].value,
                         'est': row[12].value,
                         'ger': row[15].value,
                         'ind': row[18].value,
                         'lav': row[21].value,
                         'pol': row[24].value,
                         'slo': row[27].value,
                         'swe': row[30].value}
            elif i == 1:
                i = 2
                entry['eng'] = row[4].value
                entry['chi'] = row[7].value
                entry['dan'] = row[10].value
                entry['fin'] = row[13].value
                entry['hun'] = row[16].value
                entry['ita'] = row[19].value
                entry['lit'] = row[22].value
                entry['por'] = row[25].value
                entry['slv'] = row[28].value
            elif i == 2:
                i = 0
                entry['lzh'] = row[8].value
                entry['dut'] = row[11].value
                entry['fre'] = row[14].value
                entry['ice'] = row[17].value
                entry['jpn'] = row[20].value
                entry['nno'] = row[23].value
                entry['rus'] = row[26].value
                entry['spa'] = row[29].value
                self.taxonomy[name] = entry


class IocComplementaryFile (object):
    """Represents a IOC Complementary file (Excel)."""

    def __init__(self, workbook, path):
        """Initialize with the Excel 'workbook'."""
        self.order = 4
        self.workbook = workbook
        self.path = path
        self.version = None
        self.taxonomy = {}          # This object contains taxa indexed by their name
        self.taxonomy_stats = {}
        if "IOC" in self.workbook.worksheets[0].title:
            s = self.workbook.worksheets[0].title
            regexp = re.compile("[A-Za-z ]*(\d*\.\d*).*")
            self.version = regexp.match (s).groups()[0]
        else:
            print("Error; '%s' is not a valid IOC Complementary File" % (self.path))
            raise InvalidIocComplementaryFile(self.path)

    def read(self):
        """Read the taxonomy data into the attribute 'self.taxonomy' and
           save statistics in the attribute 'self.taxonomy_stats'."""
        # Note that the global variable 'master_taxonomy' is a list that will maintain the
        # ordering of taxa in the IOC Master file.
        self.taxonomy_stats = {'infraclass_count': 0,
                               'order_count': 0,
                               'family_count': 0,
                               'genus_count': 0,
                               'species_count': 0,
                               'subspecies_count': 0}
        ws = self.workbook.worksheets[0]
        for row in ws.iter_rows(min_row=3):
            if row[1].value == "Blank":
                name = row[5].value
                self.taxonomy[name] = {'name': name,
                                       'rank': 'Infraclass',
                                       'code': row[9].value,
                                       'comment': row[10].value}
            elif row[1].value == "ORDER":
                name = row[5].value.split()[1],
                self.taxonomy[name] = {'name': name,
                                       'rank': 'Order',
                                       'code': row[9].value,
                                       'comment': row[10].value}
            elif row[1].value == "Family":
                name = row[5].value.split()[1]
                self.taxonomy[name] = {'species_count': row[2].value,
                                       'name_eng': row[3].value,
                                       'name': name,
                                       'rank': 'Family',
                                       'code': row[9].value,
                                       'comment': row[10].value}
            elif row[1].value == "Genus":
                name = row[5].value
                self.taxonomy[name] = {'extinct': row[2].value,
                                       'name': name,
                                       'rank': 'Genus',
                                       'authority': row[6].value,
                                       'code': row[9].value,
                                       'comment': row[10].value}
            elif row[1].value == "Species":
                name = row[5].value
                self.taxonomy[name] = {'extinct': row[2].value,
                                       'name': name,
                                       'rank': 'Species',
                                       'name_eng': row[3].value,
                                       'authority': row[6].value,
                                       'breeding_range': row[7].value,
                                       'nonbreeding_range': row[8].value,
                                       'code': row[9].value,
                                       'comment': row[10].value}
                species = row[6].value
            elif row[2].value == "ssp":
                name = species + ' ' + row[5].value.split()[2]
                self.taxonomy[name] = {'extinct': row[2].value,
                                       'name': name,
                                       'rank': 'Subspecies',
                                       'name_eng': row[3].value,
                                       'authority': row[6].value,
                                       'breeding_range': row[7].value,
                                       'nonbreeding_range': row[8].value,
                                       'code': row[9].value,
                                       'comment': row[10].value}


class IocData (object):
    """Represents all IOC data and the entire IOC taxonomic hierarchy. The top
       level taxa are available in the attribute 'top_taxa'. ALl other taxa can
       be retrieved by name with the feature 'taxon'."""

    def __init__(self, directory):
        """Initialize with the specified 'directory'."""
        assert os.path.exists(directory)
        assert os.path.isdir(directory)
        self.directory = directory
        self.top_taxa = self.__top_taxa__()

    def __top_taxa__(self):
        """A list of the top IocTaxon in IOC. These IocTaxa have the rank
           "Infraclass"."""
        result = []
        for name in TOP_TAXA_NAMES:
            result.append(IocTaxon(name))
        return result

    def taxon(self, name):
        """Return the IocTaxon with the given 'name'. Returns None if there is
           no such IocTaxon."""
        t = IocTaxon(name, self.directory)
        if t.exists:
            return t
        else:
            return None


class IocTaxon (object):
    """Represents an IOC taxon as transformed from the Excel data in the IOC
       Master file, Complementary file, Multilingual file and Other Lists
       file. All data for the taxon is available in the attribute 'data'."""

    def __init__(self, name, directory):
        """Initialize with the name, that can be a binomial species or a
           trinomial subspecies name."""
        self.data = {'name': name}
        self.filename = os.path.join(directory, name + '.json')
        self.exists = os.path.exists(self.filename)
        self.__load__()

    def __load__(self):
        """Load from JSON file."""
        if self.exists:
            fname = self.data['name'] + '.json'
            f = open(fname)
            self.data = json.load(f)
            f.close()
