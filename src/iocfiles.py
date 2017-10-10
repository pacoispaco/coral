#!/usr/bin/env python3

"""This module contains methods and classes for reading IOC files."""

import os, os.path
import openpyxl.reader.excel as xlxs
import re

class InvalidIocMasterFile (Exception):
    pass

class InvalidIocOtherListsFile (Exception):
    pass

class InvalidIocMultilingualFile (Exception):
    pass

class InvalidIocComplementaryFile (Exception):
    pass

class UnrecognizedIocFile (Exception):
    pass

class MultipleIocFilesOfSameType (Exception):
    pass

class MultipleVersionsOfIocFiles (Exception):
    pass

def sorted_ioc_files (filepaths):
    """Returns a list of IOC File class instances sorted in the order they
       should be processed."""
    result = []
    for path in filepaths:
        try:
            wb = xlxs.load_workbook (path)
        except xlxs.InvalidFileException:
            print ("Error: '%s' is not an Excel file (.xlxs)." % (path))
            raise
        if wb.worksheets[0].title == "Master":
            result.append (IocMasterFile (wb, path))
        elif "vs_other_lists" in wb.worksheets[0].title:
            result.append (IocOtherListsFile (wb, path))
        elif wb.worksheets[0].title == "List" and wb.worksheets[1].title == "Sources":
            result.append (IocMultilingualFile (wb, path))
        elif "IOC" in wb.worksheets[0].title:
            result.append (IocComplementaryFile (wb, path))
        else:
            print ("Error: '%s' is not an recognized IOC file." % (path))
            raise UnrecognizedIocFile (path)
    # Make sure we don't have multiple IOC files of the same type
    orders = [file.order for file in result]
    if not len (set (orders)) == len (orders):
        print ("Error: Multiple IOC files of same type.")
        raise MultipleIocFilesOfSameType (filepaths)
    # Make sure all IOC files have the same version
    versions = [file.version for file in result]
    if not len (set (versions)) == 1:
        print ("Error: Multiple versions of IOC files.")
        raise MultipleVersionsOfIocFiles (filepaths)
    # Return the list of IOC files sorted by their 'order' attribute
    return sorted (result, key=lambda iocfile: iocfile.order)

class IocMasterFile (object):
    """Represents a IOC Master file."""

    def __init__ (self, workbook, path):
        """Initialize with Excel 'workbook'."""
        self.order = 1
        self.workbook = workbook
        self.path = path
        self.version = None
        self.taxonomy = []        # This list contains the list of top level
                                  # taxa. Each taxa has a 'subtaxa' attribute
                                  # which is a list of taxa, etc.
        self.taxonomy_stats = {}
        if self.workbook.worksheets[0].title == "Master":
            s = self.workbook.worksheets[0].cell (row=1, column=2).value
            regexp = re.compile (".*\((.*)\).*")
            self.version = regexp.match (s).groups()[0]
        else:
            print ("Error; '%s' is not a avalid IOC Master File" % (filepath))
            raise InvalidIocMasterFile (filepath)

    def read (self):
        """Read the taxonomy data into the attribute 'self.taxonomy' and
           save statistics in the attribute 'self.taxonomy_stats'."""
        # Note that the global variable 'master_taxonomy' is a list that will maintain the
        # ordering of taxa in the IOC Master file.
        self.taxonomy_stats = {'infraclass_count': 0,
                               'order_count': 0,
                               'family_count': 0,
                               'genus_count': 0,
                               'species_count': 0,
                               'subspecies_count': 0}
        ws = self.workbook.worksheets[0]
        for row in ws.iter_rows (min_row=5):
            infraclass = row[0].value
            order = row[1].value
            family = row[2].value
            family_en = row[3].value
            genus = row[4].value
            species = row[5].value
            subspecies = row[6].value
            taxon = {'authority': row[7].value,
                     'names': {'en': {'ioc': row[8].value}},
                     'breeding_range': row[9].value,
                     'breeding_subranges': row[10].value,
                     'nonbreeding_range': row[11].value,
                     'code': row[12].value,
                     'comment': row[13].value,
                     'subtaxa': []}
            if infraclass:
                taxon['rank'] = "Infraclass"
                taxon['name'] = infraclass
                self.taxonomy.append (taxon)
                self.taxonomy_stats['infraclass_count'] += 1
            elif order:
                taxon['rank'] = "Order"
                taxon['name'] = order
                taxon['supertaxon'] = self.taxonomy[-1]['name']
                self.taxonomy[-1]['subtaxa'].append (taxon)
                self.taxonomy_stats['order_count'] += 1
            elif family:
                taxon['rank'] = "Family"
                taxon['name'] = family
                taxon['supertaxon'] = self.taxonomy[-1]['subtaxa'][-1]['name']
                self.taxonomy[-1]['subtaxa'][-1]['subtaxa'].append (taxon)
                self.taxonomy_stats['family_count'] += 1
            elif genus:
                taxon['rank'] = "Genus"
                taxon['name'] = genus.title ()
                taxon['supertaxon'] = self.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['name']
                self.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'].append (taxon)
                self.taxonomy_stats['genus_count'] += 1
            elif species:
                taxon['rank'] = "Species"
                taxon['name'] = species
                taxon['supertaxon'] = self.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]['name']
                genus = taxon['supertaxon']
                taxon['binomial_name'] = genus + " " + species
                self.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'].append (taxon)
                self.taxonomy_stats['species_count'] += 1
            elif subspecies:
                taxon['rank'] = "Subspecies"
                taxon['name'] = subspecies
                taxon['supertaxon'] = self.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]['name']
                genus = self.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]['name']
                species = taxon['supertaxon']
                taxon['trinomial_name'] = genus + " " + species + " " + subspecies
                self.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'].append (taxon)
                self.taxonomy_stats['subspecies_count'] += 1

class IocOtherListsFile (object):
    """Represents a IOC Other Lists file."""

    def __init__ (self, workbook, path):
        """Initialize with the Excel 'workbook'."""
        self.order = 2
        self.workbook = workbook
        self.path = path
        self.version = None
        self.taxonomy = {}          # This object contains taxa indexed by their
                                    # name (or binomial or trinomial name if
                                    # applicable (for species and subspecies)
        self.taxonomy_stats = {}
        self.lists = {'ioc_7_3': 'Gill, F & D Donsker (Eds). 2017. IOC World Bird List (v 7.3)',
                      'clements_2016': 'Clements, J. F., T. S. Schulenberg, M. J. Iliff, D. Roberson, T. A. Fredericks, B. L. Sullivan, and C. L. Wood. 2016. The eBird/Clements checklist of birds of the world: v2016.',
                      'hbwbl_2016': 'del Hoyo, J., N. J. Collar, D. A. Christie, A. Elliott,  L. D. C. Fishpool, P. Boesman & G. M. Kirwan. 2014-2016. HBW and BirdLife International Illustrated Checklist of the Birds of the World, Volume 1, 2, Lynx Edicions in association with BirdLife International, Barcelona, Spain and Cambridge, UK.',
                      'hm4_4ed': 'Dickinson, E.C., J.V. Remsen Jr. & L. Christidis (Eds). 2013-2014. The Howard & Moore Complete Checklist of the Birds of the World. 4th. Edition, Vol. 1, 2, Aves Press, Eastbourne, U.K. Plus CD content: Tyrberg, T. & E.C. Dickinson Appendix 5: Extinct species. (Also includes Errata and Corrigenda to Volume 1 and List of Errata for Vol. 2 plus Corrigenda in respect of range statements and additional Errata from Vol. 1)',
                      'hbw_2013': 'del Hoyo, J., A. Elliott, J. Sargatal & D. A. Christie (Eds). 1992-2013. Handbook of the Birds of the World, Lynx Edicions.',
                      'peters_1986': 'Peters, J.L. et al. Check-list of Birds of the World, 1931-1986. Harvard University Press/Museum of Comparative Zoology.',
                      'boyd_3_08': 'John H. Boyd III - TiF checklist, Version 3.08: May 1 2017 and updated July 12 2017',
                      'hbwbl_2017': 'BirdLife International (2017) Handbook of the Birds of the World and BirdLife International digital checklist of the birds of the world. Version 9.1. Available at: http://datazone.birdlife.org/userfiles/file/Species/Taxonomy/BirdLife_Checklist_Version_91.zip [.xls zipped 1 MB].',
                      'sibley_1993': 'Sibley, C. G. and B. L. Monroe. 1993. A Supplement to Distribution and Taxonomy of Birds of the World. Yale University Press, New Haven, Connecticut.',
                      'ioc_7_2': 'Gill, F & D Donsker (Eds). 2017. IOC World Bird List (v 7.2)',
                      'ioc_7_1': 'Gill, F & D Donsker (Eds). 2017. IOC World Bird List (v 7.1)'}
        if "vs_other_lists" in self.workbook.worksheets[0].title:
            s = self.workbook.worksheets[0].cell (row=1, column=2).value
            regexp = re.compile (".*\(v (.*)\).*")
            self.version = regexp.match (s).groups()[0]
        else:
            print ("Error; '%s' is not a valid IOC Other Lists File" % (filepath))
            raise InvalidIocOtherListsFile (filepath)

    def read (self):
        """Read the taxonomy data into the attribute 'self.taxonomy' and
           save statistics in the attribute 'self.taxonomy_stats'."""
        # Note that the global variable 'master_taxonomy' is a list that will maintain the
        # ordering of taxa in the IOC Master file.
        self.taxonomy_stats = {'infraclass_count': 0,
                               'order_count': 0,
                               'family_count': 0,
                               'genus_count': 0,
                               'species_count': 0,
                               'subspecies_count': 0,
                               'only_in_other_lists_count': 0}
        ws = self.workbook.worksheets[0]
        for row in ws.iter_rows (min_row=2):
            name = row[1].value
            self.taxonomy[name] = {'seq_no': row[0].value,
                                   'name': name,
                                   'notes': row[2].value,
                                   'iucn_red_list_category': row[27].value,
                                   'lists': {'clements_2016': {'name': row[4].value,
                                                               'group': row[3].value,
                                                               'family': row[17].value},
                                             'hbwbl_2016': {'name': row[5].value,
                                                            'group': row[6].value,
                                                            'family': row[18].value},
                                             'hm4_4ed': {'name': row[8].value,
                                                         'group': row[7].value,
                                                         'family': row[19].value},
                                             'hbw_2013': {'name': row[9].value,
                                                          'group': None,
                                                          'family': row[20].value},
                                             'peters_1986': {'name': row[10].value,
                                                             'group': None,
                                                             'family': row[21].value},
                                             'boyd': {'name': row[11].value,
                                                      'group': None,
                                                      'family': row[22].value},
                                             'hbwbl_2017': {'name': row[12].value,
                                                            'group': None,
                                                            'family': row[23].value},
                                             'sibley_1993': {'name': row[13].value,
                                                             'group': None,
                                                             'family': row[24].value},
                                             'ioc_7_2': {'name': None,
                                                         'group': None,
                                                         'family': row[25].value},
                                             'ioc_7_1': {'name': None,
                                                         'group': None,
                                                         'family': row[26].value}}}
            if name:
                if len (name.split ()) == 2:
                    self.taxonomy_stats['species_count'] += 1
                else:
                    self.taxonomy_stats['subspecies_count'] += 1
            else:
                self.taxonomy_stats['only_in_other_lists_count'] += 1

class IocMultilingualFile (object):
    """Represents a IOC Multilingual file. Languages are encoded with ISO 639-2
       codes."""

    def __init__ (self, workbook, path):
        """Initialize with the Excel 'workbook'."""
        self.order = 3
        self.workbook = workbook
        self.path = path
        self.version = None
        self.taxonomy = {}          # This object contains taxa indexed by their
                                    # name
        self.taxonomy_stats = {}
        if self.workbook.worksheets[0].title == "List" and self.workbook.worksheets[1].title == "Sources":
            s = self.workbook.worksheets[1].cell (row=1, column=1).value
            regexp = re.compile ("[A-Za-z ]*(\d*\.\d*).*")
            self.version = regexp.match (s).groups()[0]
        else:
            print ("Error; '%s' is not a valid IOC Multilingual File" % (filepath))
            raise InvalidIocMultilingualFile (filepath)

    def read (self):
        """Read the taxonomy data into the attribute 'self.taxonomy' and
           save statistics in the attribute 'self.taxonomy_stats'."""
        # Note that the global variable 'master_taxonomy' is a list that will maintain the
        # ordering of taxa in the IOC Master file.
        self.taxonomy_stats = {'infraclass_count': 0,
                               'order_count': 0,
                               'family_count': 0,
                               'genus_count': 0,
                               'species_count': 0,
                               'subspecies_count': 0}
        ws = self.workbook.worksheets[0]
        i = 0
        for row in ws.iter_rows (min_row=4):
            if row[3].value and len (row[3].value.split ()) == 2:
                name = row[3].value
                self.taxonomy_stats['species_count'] += 1
                i = 1
                entry = {'name': name,
                         'cat': row[6].value,
                         'cze': row[9].value,
                         'est': row[12].value,
                         'ger': row[15].value,
                         'ind': row[18].value,
                         'lav': row[21].value,
                         'pol': row[24].value,
                         'slo': row[27].value,
                         'swe': row[30].value}
            elif i == 1:
                i = 2
                entry['eng'] = row[4].value
                entry['chi'] = row[7].value
                entry['dan'] = row[10].value
                entry['fin'] = row[13].value
                entry['hun'] = row[16].value
                entry['ita'] = row[19].value
                entry['lit'] = row[22].value
                entry['por'] = row[25].value
                entry['slv'] = row[28].value
            elif i == 2:
                i = 0
                entry['lzh'] = row[8].value
                entry['dut'] = row[11].value
                entry['fre'] = row[14].value
                entry['ice'] = row[17].value
                entry['jpn'] = row[20].value
                entry['nno'] = row[23].value
                entry['rus'] = row[26].value
                entry['spa'] = row[29].value
                self.taxonomy[name] = entry

class IocComplementaryFile (object):
    """Represents a IOC COmplementary file."""

    def __init__ (self, workbook, path):
        """Initialize with the Excel 'workbook'."""
        self.order = 4
        self.workbook = workbook
        self.path = path
        self.version = None
        self.taxonomy = {}          # This object contains taxa indexed by their
                                    # name
        self.taxonomy_stats = {}
        if "IOC" in self.workbook.worksheets[0].title:
            s = self.workbook.worksheets[0].title
            regexp = re.compile ("[A-Za-z ]*(\d*\.\d*).*")
            self.version = regexp.match (s).groups()[0]
        else:
            print ("Error; '%s' is not a valid IOC Complementary File" % (filepath))
            raise InvalidIocComplemntaryFile (filepath)

    def read (self):
        """Read the taxonomy data into the attribute 'self.taxonomy' and
           save statistics in the attribute 'self.taxonomy_stats'."""
        # Note that the global variable 'master_taxonomy' is a list that will maintain the
        # ordering of taxa in the IOC Master file.
        self.taxonomy_stats = {'infraclass_count': 0,
                               'order_count': 0,
                               'family_count': 0,
                               'genus_count': 0,
                               'species_count': 0,
                               'subspecies_count': 0}
        ws = self.workbook.worksheets[0]
        i = 0
        for row in ws.iter_rows (min_row=3):
            if row[2].value == "Blank":
                name = row[6].value
                self.taxonomy[name] = {'name': name,
                                       'rank': 'Infraclass',
                                       'code': row[10].value,
                                       'comment': row[11].value}
            elif row[2].value == "ORDER":
                name = row[6].value.split ()[1],
                self.taxonomy[name] = {'name': name,
                                       'rank': 'Order',
                                       'code': row[10].value,
                                       'comment': row[11].value}
            elif row[2].value == "Family":
                name = row[6].value.split ()[1]
                self.taxonomy[name] = {'species_count': row[3].value,
                                       'name_eng': row[4].value,
                                       'name': name,
                                       'rank': 'Family',
                                       'code': row[10].value,
                                       'comment': row[11].value}
            elif row[2].value == "Genus":
                name = row[6].value
                self.taxonomy[name] = {'name': name,
                                       'rank': 'Genus',
                                       'authority': row[7].value,
                                       'code': row[10].value,
                                       'comment': row[11].value}
            elif row[2].value == "Species":
                name = row[6].value
                self.taxonomy[name] = {'name': name,
                                       'rank': 'Species',
                                       'name_eng': row[4].value,
                                       'authority': row[7].value,
                                       'breeding_range': row[8].value,
                                       'nonbreeding_range': row[9].value,
                                       'code': row[10].value,
                                       'comment': row[11].value}
                species = row[6].value
            elif row[2].value == "ssp":
                name = species + ' ' + row[6].value.split ()[2]
                self.taxonomy[name] = {'extinct': row[3].value,
                                       'name': name,
                                       'rank': 'Subspecies',
                                       'name_eng': row[4].value,
                                       'authority': row[7].value,
                                       'breeding_range': row[8].value,
                                       'nonbreeding_range': row[9].value,
                                       'code': row[10].value,
                                       'comment': row[11].value}
