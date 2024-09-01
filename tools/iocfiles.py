#!/usr/bin/env python3

"""This module contains methods and classes for reading IOC files."""

import json
import os
import os.path
import openpyxl.reader.excel as xlxs
import re

# Constants: top taxa names in IOC
TOP_TAXA_NAMES = ["NEOAVES", "NEOGNATHAE", "PALEOGNATHAE"]

# Directory and file name constants
DEFAULT_DATA_DIR = "./gendata"
DEFAULT_IOC_TAXONOMY_DIR = "ioc"
VERSION_FILE_NAME = "version.json"


class InvalidIocMasterFile (Exception):
    pass


class InvalidIocOtherListsFile (Exception):
    pass


class InvalidIocMultilingualFile (Exception):
    pass


class InvalidIocComplementaryFile (Exception):
    pass


class UnrecognizedIocFile (Exception):
    pass


class MultipleIocFilesOfSameType (Exception):
    pass


class MultipleVersionsOfIocFiles (Exception):
    pass


class IocWblDirectoryNotFound(Exception):
    pass


class IocWbl(object):
    """Representation of IOC World Bird List"""

    def __init__(self, dirpath=None):
        """Initailize the IOC World Bird List. If `dirpath`is None, and empty object will be
           created. If `dirpath` is not None, it will try to read the IOC World Bird List
           JSON-files from there. That assumes that the IOC Name List file has been read and
           the corresponding IOC World Bird List JSON-files have been created at an earlier
           point in time."""
        self.taxonomy = []
        self.index = {}
        self.version = None
        self.stats = {'infraclass_count': 0,
                      'order_count': 0,
                      'family_count': 0,
                      'genus_count': 0,
                      'species_count': 0,
                      'subspecies_count': 0}
        if dirpath:
            if os.path.exists(dirpath):
                self.load_taxonomy(dirpath, version_file_name=VERSION_FILE_NAME)
            else:
                print(f"Error: IOC World Bird List directory {dirpath} not found.")
                raise IocWblDirectoryNotFound(dirpath)

    def _write_taxon_to_file(self, directory, taxon):
        """Write the JSON representation of 'taxon' to file."""
        for subtaxon in taxon['subtaxa']:
            self._write_taxon_to_file(directory, subtaxon)
        if taxon['rank'] == "Species":
            fname = taxon['binomial_name'].replace(" ", "_") + ".json"
        elif taxon['rank'] == "Subspecies":
            fname = taxon['trinomial_name'].replace(" ", "_") + ".json"
        else:
            fname = taxon['name'] + ".json"
        subtaxa_files = []
        for subtaxon in taxon['subtaxa']:
            if subtaxon['rank'] == "Species":
                subtaxa_files.append(subtaxon['binomial_name'].replace(" ", "_") + ".json")
            elif subtaxon['rank'] == "Subspecies":
                subtaxa_files.append(subtaxon['trinomial_name'].replace(" ", "_") + ".json")
            else:
                subtaxa_files.append(subtaxon['name'] + ".json")
        taxon['subtaxa'] = subtaxa_files
        f = open(os.path.join(directory, fname), 'w')
        f.write(json.dumps(taxon))
        f.close()

    def write_to_files(self, dirpath, version_file_name=VERSION_FILE_NAME):
        """Write JSON representations of the taxa in IOC taxonomy to files in the directory
           `dirpath`."""
        assert os.path.exists(dirpath)
        f = open(os.path.join(dirpath, version_file_name), 'w')
        v = {"version": self.version}
        f.write(json.dumps(v))
        f.close()
        for taxon in self.taxonomy:
            self._write_taxon_to_file(dirpath, taxon)

    def _load_subtaxa(self, taxon, dirpath):
        """Load the subtaxa for the given taxon."""
        i = 0
        for name in taxon['subtaxa']:
            f = open(os.path.join(dirpath, name))
            t = json.load(f)
            f.close()
            taxon['subtaxa'][i] = t
            if t['rank'] == "Order":
                self.index[t['name']] = t
                self.stats['order_count'] += 1
            elif t['rank'] == "Family":
                self.index[t['name']] = t
                self.stats['family_count'] += 1
            elif t['rank'] == "Genus":
                self.index[t['name']] = t
                self.stats['genus_count'] += 1
            elif t['rank'] == "Species":
                self.index[t['binomial_name']] = t
                self.stats['species_count'] += 1
            elif t['rank'] == "Subspecies":
                self.index[t['trinomial_name']] = t
                self.stats['subspecies_count'] += 1
            self._load_subtaxa(t, dirpath)
            i += 1

    def load_taxonomy(self, dirpath, version_file_name=VERSION_FILE_NAME):
        """Load IOC taxonomy from files in the directory `dirpath`."""
        # Read version
        p = os.path.join(DEFAULT_DATA_DIR, DEFAULT_IOC_TAXONOMY_DIR)
        f = open(os.path.join(p, VERSION_FILE_NAME))
        self.version = json.load(f)["version"]
        f.close()
        # Read infraclasses:
        p = os.popen("grep -l '\"rank\": \"Infraclass\"' %s/*.json" % (dirpath))
        filenames = p.read().split()
        p.close()
        for fname in filenames:
            f = open(fname)
            taxon = json.load(f)
            f.close()
            self.index[taxon['name']] = taxon
            self.taxonomy.append(taxon)
            self._load_subtaxa(taxon, dirpath)
            self.stats['infraclass_count'] += 1


class IocMasterFile (object):
    """Represents a IOC Master file (Excel). This is the first file that must be read. It will set
       up an IocWbl object that can then be passed to the other 3 IOC file clases."""

    def __init__(self, filepath):
        """Initialize with Excel file `filepath`. This will not read the contents of the file."""
        # Check that it is an Excel-file.
        try:
            wb = xlxs.load_workbook(filepath)
        except xlxs.InvalidFileException:
            print(f"Error: {filepath} is not an Excel file (.xlxs).")
            raise
        # Check that it is a IOC Master file, and if so initialize it.
        if self._is_master_wb(wb):
            self.order = 1
            self.workbook = wb
            self.path = filepath
            self.version = self._master_wb_version(self.workbook)
            self.iocwbl = IocWbl()
        else:
            print((f"Error: '{filepath}' is not a valid IOC Master File.\n"
                   f"An IOC Master file must have the title 'Master' in the first worksheet."))
            raise InvalidIocMasterFile(self.path)
        # Check the version, to see if we need to do column shift when reading data from it.
        if self.version in ["8.1", "7.3"]:
            self.column_shift = 0
        elif self.version >= "14.1":
            self.column_shift = 1

    def _is_master_wb(self, wb):
        """True if 'wb' is an IOC Masterfile Excel workbook, otherwise False."""
        return wb.worksheets[0].title == "Master"

    def _master_wb_version(self, wb):
        """Returns the IOC version number of the workbook. Returns None if not able to establish
           the version."""
        if self._is_master_wb(wb):
            if wb.worksheets[0].cell(row=1, column=2).value:
                version_string = wb.worksheets[0].cell(row=1, column=2).value
            elif wb.worksheets[0].cell(row=1, column=3).value:
                version_string = wb.worksheets[0].cell(row=1, column=3).value
            regexp = re.compile(r"IOC WORLD BIRD LIST \((.*)\)")
            mo = regexp.search(version_string)
            return mo[1]
        else:
            return None

    def read(self):
        """Read the taxonomy data into the attribute 'self.taxonomy' and
           save statistics in the attribute 'self.taxonomy_stats'."""
        self.taxonomy_stats = {'infraclass_count': 0,
                               'order_count': 0,
                               'family_count': 0,
                               'genus_count': 0,
                               'species_count': 0,
                               'subspecies_count': 0}
        ws = self.workbook.worksheets[0]
        for row in ws.iter_rows(min_row=5):
            infraclass = row[0].value
            order = row[1+self.column_shift].value
            family = row[2+self.column_shift].value
            family_en = row[3+self.column_shift].value
            genus = row[4+self.column_shift].value
            species = row[5+self.column_shift].value
            subspecies = row[6+self.column_shift].value
            taxon = {'other_classifications': [],  # Taxa in other lists which
                                                   # are equivalent to this taxon
                     'authority': row[7+self.column_shift].value,
                     'common_names': {'en': row[8+self.column_shift].value},
                     'breeding_range': row[9+self.column_shift].value,
                     'breeding_subranges': row[10+self.column_shift].value,
                     'nonbreeding_range': row[11+self.column_shift].value,
                     'code': row[12+self.column_shift].value,
                     'comment': row[13+self.column_shift].value,
                     'subtaxa': []}
            if infraclass:
                taxon['rank'] = "Infraclass"
                taxon['name'] = infraclass
                self.iocwbl.taxonomy.append(taxon)
                self.iocwbl.index[infraclass] = taxon
                self.iocwbl.stats['infraclass_count'] += 1
            elif order:
                taxon['rank'] = "Order"
                taxon['name'] = order
                taxon['supertaxon'] = self.iocwbl.taxonomy[-1]['name']
                self.iocwbl.taxonomy[-1]['subtaxa'].append(taxon)
                self.iocwbl.index[order] = taxon
                self.iocwbl.stats['order_count'] += 1
            elif family:
                taxon['rank'] = "Family"
                taxon['name'] = family
                taxon['common_names'] = {'en': family_en}
                taxon['supertaxon'] = self.iocwbl.taxonomy[-1]['subtaxa'][-1]['name']
                self.iocwbl.taxonomy[-1]['subtaxa'][-1]['subtaxa'].append(taxon)
                self.iocwbl.index[family] = taxon
                self.iocwbl.stats['family_count'] += 1
            elif genus:
                taxon['rank'] = "Genus"
                # Strip trailing "extinct" characters '\u2020' and whitespace
                taxon['name'] = genus.title().strip('\u2020').strip()
                taxon['supertaxon'] = self.iocwbl.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['name']
                self.iocwbl.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'].append(taxon)
                self.iocwbl.index[genus] = taxon
                self.iocwbl.stats['genus_count'] += 1
            elif species:
                taxon['rank'] = "Species"
                taxon['name'] = species
                taxon['supertaxon'] = (self.iocwbl.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]
                                       ['subtaxa'][-1]['name'])
                genus = taxon['supertaxon']
                # Strip trailing "extinct" characters '\u2020' and whitespace
                taxon['binomial_name'] = (genus + " " + species).strip('\u2020').strip()
                subtaxa = (self.iocwbl.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]
                           ['subtaxa'])
                subtaxa.append(taxon)
                self.iocwbl.index[taxon['binomial_name']] = taxon
                self.iocwbl.stats['species_count'] += 1
            elif subspecies:
                taxon['rank'] = "Subspecies"
                taxon['name'] = subspecies
                taxon['supertaxon'] = (self.iocwbl.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]
                                       ['subtaxa'][-1]['subtaxa'][-1]['name'])
                genus = (self.iocwbl.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]
                         ['name'])
                species = taxon['supertaxon']
                # Strip trailing "extinct" characters '\u2020' and whitespace
                s = genus + " " + species + " " + subspecies
                trinomial_name = s.strip('\u2020').strip()
                # Strip "extinct" characters '\u2020' in trinomial name
                taxon['trinomial_name'] = trinomial_name.replace(" \u2020", "")
                subtaxa = (self.iocwbl.taxonomy[-1]['subtaxa'][-1]['subtaxa'][-1]['subtaxa'][-1]
                           ['subtaxa'][-1]['subtaxa'])
                subtaxa.append(taxon)
                self.iocwbl.index[taxon['trinomial_name']] = taxon
                self.iocwbl.stats['subspecies_count'] += 1


class IocOtherListsFile (object):
    """Represents a IOC Other Lists file (Excel). NOTE: This file seems to have been dropped from
       the IOC files somewhere between version 8.1 and 14.1."""

    def __init__(self, filepath, iocwbl):
        """Initialize with Excel file `filepath` and an `iocwbl` object obtained from an
           `IocMasterFile` object. This will not read the contents of the file."""
        # Check that it is an Excel-file.
        try:
            wb = xlxs.load_workbook(filepath)
        except xlxs.InvalidFileException:
            print(f"Error: {filepath} is not an Excel file (.xlxs).")
            raise
        # Check that it is a IOC Other Lists file, and if so initialize it.
        if self._is_other_lists_wb(wb):
            self.iocwbl = iocwbl
            self.order = 2
            self.path = filepath
            self.workbook = wb
            self.version = self._other_lists_wb_version(wb)
            self.taxonomy_stats = {}
            self.lists = {
                'ioc_7_3': 'Gill, F & D Donsker (Eds). 2017. IOC World Bird List (v 7.3)',
                'clements_2016': (
                    "Clements, J. F., T. S. Schulenberg, M. J. Iliff, D. Roberson, T. A. "
                    "Fredericks, B. L. Sullivan, 'C. L. Wood, and D. F. A. Wiedenfeld. 2016. "
                    "The Clements Checklist of Birds of the World: Version 6.9. Cornell Lab of "
                    "Ornithology."),
                'hbwbl_2016': (
                    "del Hoyo, J., N. J. Collar, D. A. Christie, A. Elliott, L. D. C. Fishpool, "
                    "P. Boesman & G. M. Kirwan (Eds). 2016. Handbook of the Birds of the World "
                    "and BirdLife International Illustrated Checklist of the Birds of the World. "
                    "Volume 1: Non-Passerines."),
                'hm4_4ed': (
                    "Dickinson, E.C., J.V. Remsen Jr. & L. Christidis (Eds). 2013-2014. The "
                    "Howard & Moore Complete Checklist of the Birds of the World. 4th Edition. "
                    "Aves Press."),
                'hbw_2013': (
                    "del Hoyo, J., A. Elliott, J. Sargatal & D. A. Christie (Eds). 1992-2013. "
                    "Handbook of the Birds of the World. Vols. 1-19. Lynx Edicions."),
                'peters_1986': (
                    "Peters, J.L. et al. Check-list of Birds of the World, 1931-1986. Harvard "
                    "University Press/Museum of Comparative Zoology."),
                'boyd_3_08': (
                    "John H. Boyd III - TiF checklist, Version 3.08: May 1 2017 and updated July "
                    "12 2017."),
                'hbwbl_2017': (
                    "BirdLife International (2017) Handbook of the Birds of the World and "
                    "BirdLife International digital checklist. Version 1.0."),
                'sibley_1993': (
                    "Sibley, C. G. and B. L. Monroe. 1993. A Supplement to Distribution and "
                    "Taxonomy of Birds of the World."),
                'ioc_7_2': 'Gill, F & D Donsker (Eds). 2017. IOC World Bird List (v 7.2)',
                'ioc_7_1': 'Gill, F & D Donsker (Eds). 2017. IOC World Bird List (v 7.1)'
            }
        else:
            print((f"Error: '{filepath}' is not a valid IOC Other Lists File.\nAn IOC Other "
                   "Lists file must have the word 'vs_other_lists' in the title of the first "
                   "worksheet."))
            raise InvalidIocOtherListsFile(self.path)

    def _is_other_lists_wb(self, wb):
        """True if 'wb' is an IOC Other Lists Excel workbook, otherwise False."""
        return "vs_other_lists" in wb.worksheets[0].title

    def _other_lists_wb_version(self, wb):
        """Returns the IOC version number of the workbook. Returns None if not able to establish
           the version."""
        if self._is_other_lists_wb(wb):
            s = wb.worksheets[0].cell(row=1, column=2).value
            regexp = re.compile(r".*\(v (.*)\).*")
            self.version = regexp.match(s).groups()[0]
        else:
            return None

    def read(self):
        """Read the taxonomy data into the attribute 'self.taxonomy' and
           save statistics in the attribute 'self.taxonomy_stats'."""
        # Note that the global variable 'master_taxonomy' is a list that will maintain the
        # ordering of taxa in the IOC Master file.
        self.taxonomy_stats = {'infraclass_count': 0,
                               'order_count': 0,
                               'family_count': 0,
                               'genus_count': 0,
                               'species_count': 0,
                               'subspecies_count': 0,
                               'only_in_other_lists_count': 0}
        ws = self.workbook.worksheets[0]
        for row in ws.iter_rows(min_row=2):
            name = row[1].value
            entry = {'seq_no': row[0].value,
                     'name': name,
                     'rank': row[2].value,
                     'notes': row[3].value,
                     'iucn_red_list_category': row[32].value,
                     'following_entries': [],
                     'lists': {'clements_2016': {'name': row[5].value,
                                                 'group': row[4].value,
                                                 'family': row[18].value},
                               'hbwbl_2016': {'name': row[7].value,
                                              'group': row[6].value,
                                              'family': row[19].value},
                               'hm4_4ed': {'name': row[9].value,
                                           'group': row[8].value,
                                           'family': row[20].value},
                               'hbw_2013': {'name': row[10].value,
                                            'group': None,
                                            'family': row[21].value},
                               'peters_1986': {'name': row[11].value,
                                               'group': None,
                                               'family': row[22].value},
                               'boyd': {'name': row[12].value,
                                        'group': None,
                                        'family': row[23].value},
                               'hbwbl_2017': {'name': row[13].value,
                                              'group': None,
                                              'family': row[24].value},
                               'sibley_1993': {'name': row[14].value,
                                               'group': None,
                                               'family': row[25].value},
                               'ioc_7_2': {'name': row[15].value,
                                           'group': None,
                                           'family': row[26].value},
                               'ioc_7_1': {'name': row[16].value,
                                           'group': None,
                                           'family': row[27].value}}}
            if name:
                self.taxonomy[name] = entry
                if len(name.split()) == 2:
                    self.taxonomy_stats['species_count'] += 1
                else:
                    self.taxonomy_stats['subspecies_count'] += 1
                latest_name = name
            else:
                self.nonindexed.append(entry)
                self.taxonomy[latest_name]['following_entries'].append(entry)
                self.taxonomy_stats['only_in_other_lists_count'] += 1

    def _add_other_lists(self):
        """Add other lists based on the IOC Other Lists File to `self.iocwbl`."""
        # TBD! We probably need to store the data we read from the iocwlb
        # as a list of "lines"/"entries" each containing information on a taxa,
        # maybe complementing it with an index of taxon names pointing to the
        # corresponding "line"/"entry".
        for name in iter(self.iocwbl.index):
            if name in self.taxonomy:
                entries = self.taxonomy[name]["following_entries"]
                lists = self.taxonomy[name]["lists"]
                self.iocwbl.index[name]["following_entries"] = (entries)
                self.iocwbl.index[name]["lists"] = (lists)


class IocMultilingualFile (object):
    """Represents a IOC Multilingual file (Excel). Languages are encoded with ISO 639-2 codes
       (which are not used in the actual file)."""

    def __init__(self, filepath, iocwbl):
        """Initialize with Excel file `filepath` and an `iocwbl` object obtained from an
           `IocMasterFile` object. This will not read the contents of the file."""
        # Check that it is an Excel-file.
        try:
            wb = xlxs.load_workbook(filepath)
        except xlxs.InvalidFileException:
            print(f"Error: {filepath} is not an Excel file (.xlxs).")
            raise
        # Check that it is a IOC Multilingual file, and if so initialize it.
        if self._is_multilingual_wb(wb):
            self.iocwbl = iocwbl
            self.order = 3
            self.workbook = wb
            self.path = filepath
            self.version = self._multilingual_wb_version(wb)
            self.taxonomy = {}          # This object contains taxa indexed by their name
            self.taxonomy_stats = {}
        else:
            print((f"Error: '{filepath}' is not a valid IOC Multilingual File.\nAn IOC Multilingual"
                   "file must have the word 'List' in the title of the first worksheet and"
                   "'Sources' in the title of the second worksheet."))
            raise InvalidIocMultilingualFile(self.path)

    def _is_multilingual_wb(self, wb):
        """True if 'wb' is an IOC Multilingual Excel workbook, otherwise False."""
        return wb.worksheets[0].title == "List" and wb.worksheets[1].title == "Sources"

    def _multilingual_wb_version(self, wb):
        """Returns the IOC version number of the workbook. Returns None if not able to establish
           the version."""
        if self._is_multilingual_wb(wb):
            if wb.worksheets[0].cell(row=1, column=4).value[0:4] == "IOC_":
                # This shouild work for "IOC_14.1" and later
                return wb.worksheets[0].cell(row=1, column=4).value[4:]
            elif wb.worksheets[0].cell(row=1, column=4).value == "Scientific Name 8.1":
                return "8.1"
            elif wb.worksheets[0].cell(row=1, column=1).value == "7.3":
                return "7.3"
            else:
                return None
        else:
            return None

    def _add_languages(self):
        """Add languages from the IOC Multilingual file to `self.iocwbl`."""
        for name in iter(self.iocwbl.index):
            if name in self.taxonomy:
                self.iocwbl.index[name]["common_names"].update(self.taxonomy[name])

    def read(self):
        """Read the taxonomy data into the attribute 'self.taxonomy' and
           save statistics in the attribute 'self.taxonomy_stats'."""
        # Note that the global variable 'master_taxonomy' is a list that will maintain the
        # ordering of taxa in the IOC Master file.
        self.taxonomy_stats = {'infraclass_count': 0,
                               'order_count': 0,
                               'family_count': 0,
                               'genus_count': 0,
                               'species_count': 0,
                               'subspecies_count': 0}
        ws = self.workbook.worksheets[0]
        if self.version in ["8.1", "7.3"]:
            i = 0
            for row in ws.iter_rows(min_row=4):
                if row[3].value and len(row[3].value.split()) == 2:
                    name = row[3].value
                    self.taxonomy_stats['species_count'] += 1
                    i = 1
                    entry = {'cat': row[6].value,
                             'cze': row[9].value,
                             'est': row[12].value,
                             'ger': row[15].value,
                             'ind': row[18].value,
                             'lav': row[21].value,
                             'pol': row[24].value,
                             'slo': row[27].value,
                             'swe': row[30].value}
                elif i == 1:
                    i = 2
                    entry['eng'] = row[4].value
                    entry['chi'] = row[7].value
                    entry['dan'] = row[10].value
                    entry['fin'] = row[13].value
                    entry['hun'] = row[16].value
                    entry['ita'] = row[19].value
                    entry['lit'] = row[22].value
                    entry['por'] = row[25].value
                    entry['slv'] = row[28].value
                elif i == 2:
                    i = 0
                    entry['lzh'] = row[8].value
                    entry['dut'] = row[11].value
                    entry['fre'] = row[14].value
                    entry['ice'] = row[17].value
                    entry['jpn'] = row[20].value
                    entry['nno'] = row[23].value
                    entry['rus'] = row[26].value
                    entry['spa'] = row[29].value
                    self.taxonomy[name] = entry
        else:  # self.version == "14.1"
            # We use IETF BCP 47 language codes. There is a Python module 'langcodes' that can be
            # used to work with them. E.g: `langcodes.get("zh-Hant").display_name()` will return
            # 'Chinese (Traditional)'.
            for row in ws.iter_rows(min_row=2):
                name = row[3].value
                entry = {'en': row[4].value,
                         'ca': row[5].value,
                         'zh-Hans': row[6].value,
                         'zh-Hant': row[7].value,
                         'hr': row[8].value,
                         'cs': row[9].value,
                         'da': row[10].value,
                         'nl': row[11].value,
                         'fi': row[12].value,
                         'fr': row[13].value,
                         'de': row[14].value,
                         'it': row[15].value,
                         'ja': row[16].value,
                         'lt': row[17].value,
                         'no': row[18].value,
                         'pl': row[19].value,
                         'pt-br': row[20].value,
                         'pt': row[21].value,
                         'ru': row[22].value,
                         'sr': row[23].value,
                         'sk': row[24].value,
                         'es': row[25].value,
                         'sv': row[26].value,
                         'tr': row[27].value,
                         'uk': row[28].value,
                         'af': row[29].value,
                         'ar': row[30].value,
                         'be': row[31].value,
                         'bg': row[32].value,
                         'et': row[33].value,
                         'el': row[34].value,
                         'he': row[35].value,
                         'hu': row[36].value,
                         'is': row[37].value,
                         'id': row[38].value,
                         'ko': row[39].value,
                         'lv': row[40].value,
                         'mk': row[41].value,
                         'ml': row[42].value,
                         'se': row[43].value,
                         'fa': row[44].value,
                         'ro': row[45].value,
                         'sl': row[46].value,
                         'th': row[47].value}
                self.taxonomy[name] = entry
                self.taxonomy_stats['species_count'] += 1
        self._add_languages()


class IocComplementaryFile (object):
    """Represents a IOC Complementary file (Excel)."""

    def __init__(self, filepath, iocwbl):
        """Initialize with Excel file `filepath`. This will not read the contents of the file."""
        # Check that it is an Excel-file.
        try:
            wb = xlxs.load_workbook(filepath)
        except xlxs.InvalidFileException:
            print(f"Error: {filepath} is not an Excel file (.xlxs).")
            raise
        # Check that it is a IOC Multilingual file, and if so initialize it.
        if self._is_complementary_wb(wb):
            self.iocwbl = iocwbl
            self.order = 4
            self.workbook = wb
            self.path = filepath
            self.version = self._complementary_wb_version(self.workbook)
            self.taxonomy = {}          # This object contains taxa indexed by their name
            self.taxonomy_stats = {}
        # Check the version, to see if we need to do column shift when reading data from it.
            if self.version in ["8.1", "7.3"]:
                self.column_shift = 0
            elif self.version in ["14.1", "14.2"]:
                self.column_shift = 1
        else:
            print((f"Error: '{filepath}' is not a valid IOC Complementary File.\nThe first "
                   "worksheet title of an IOC Complementary file must be 'IOC 7.3', 'IOC 8.1' "
                   "or '14.1'."))
            raise InvalidIocComplementaryFile(self.path)

    def _is_complementary_wb(self, wb):
        """True if 'wb' is an IOC Complementary Excel workbook, otherwise False. These files are
           normally named 'IOC_Names_File_Plus-N.M.xlxs, where N is the major version number and
           M is the minor version number."""
        sheet = wb.worksheets[0]
        version = self._complementary_wb_version(wb)
        if version:
            if version in ["7.3", "8.1"]:
                if ((sheet.cell(row=1, column=4).value == "English name") and
                   (sheet.cell(row=1, column=5).value == "Counters")):
                    return True
                else:
                    return False
            elif version in ["14.1", "14.2"]:
                if ((sheet.cell(row=1, column=6).value == "English name") and
                   (sheet.cell(row=1, column=7).value == "Counters")):
                    return True
                else:
                    return False
        else:
            return False

    def _complementary_wb_version(self, wb):
        """Returns the IOC version number of the workbook. Returns None if not able to establish
           the version."""
        if wb.worksheets[0].title == "IOC 7.3":
            return "7.3"
        elif wb.worksheets[0].title == "IOC 8.1":
            return "8.1"
        else:
            # This works for 14.1 and 14.2
            return wb.worksheets[0].title

    def _add_complementary_info(self):
        """Add complementary information from the IOC Complementary file to `self.iocwbl`."""
        for name in iter(self.iocwbl.index):
            relevant_ranks = ["Genus", "Species", "Subspecies"]
            if name in self.taxonomy and self.iocwbl.index[name]["rank"] in relevant_ranks:
                self.iocwbl.index[name]["extinct"] = self.taxonomy[name]["extinct"]
                self.iocwbl.index[name]["code"] = self.taxonomy[name]["code"]

    def read(self):
        """Read the taxonomy data into the attribute 'self.taxonomy' and
           save statistics in the attribute 'self.taxonomy_stats'."""
        # Note that the global variable 'master_taxonomy' is a list that will maintain the
        # ordering of taxa in the IOC Master file.
        self.taxonomy_stats = {'infraclass_count': 0,
                               'order_count': 0,
                               'family_count': 0,
                               'genus_count': 0,
                               'species_count': 0,
                               'subspecies_count': 0}
        ws = self.workbook.worksheets[0]
        for row in ws.iter_rows(min_row=3):
            if row[1].value == "Blank":
                name = row[5].value
                self.taxonomy[name] = {'name': name,
                                       'rank': 'Infraclass',
                                       'code': row[9].value,
                                       'comment': row[10].value}
            elif row[1].value == "ORDER":
                name = row[5].value.split()[1],
                self.taxonomy[name] = {'name': name,
                                       'rank': 'Order',
                                       'code': row[9].value,
                                       'comment': row[10].value}
            elif row[1].value == "Family":
                name = row[5].value.split()[1]
                self.taxonomy[name] = {'species_count': row[2].value,
                                       'name_eng': row[3].value,
                                       'name': name,
                                       'rank': 'Family',
                                       'code': row[9].value,
                                       'comment': row[10].value}
            elif row[1].value == "Genus":
                name = row[5].value
                self.taxonomy[name] = {'extinct': row[2].value,
                                       'name': name,
                                       'rank': 'Genus',
                                       'authority': row[6].value,
                                       'code': row[9].value,
                                       'comment': row[10].value}
            elif row[1].value == "Species":
                name = row[5].value
                self.taxonomy[name] = {'extinct': row[2].value,
                                       'name': name,
                                       'rank': 'Species',
                                       'name_eng': row[3].value,
                                       'authority': row[6].value,
                                       'breeding_range': row[7].value,
                                       'nonbreeding_range': row[8].value,
                                       'code': row[9].value,
                                       'comment': row[10].value}
                species = row[6].value
            elif row[2].value == "ssp":
                name = species + ' ' + row[5].value.split()[2]
                self.taxonomy[name] = {'extinct': row[2].value,
                                       'name': name,
                                       'rank': 'Subspecies',
                                       'name_eng': row[3].value,
                                       'authority': row[6].value,
                                       'breeding_range': row[7].value,
                                       'nonbreeding_range': row[8].value,
                                       'code': row[9].value,
                                       'comment': row[10].value}
        self._add_complementary_info()
